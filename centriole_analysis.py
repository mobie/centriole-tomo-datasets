#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Apr 22 08:52:52 2021

@author: schorb
"""

from openpyxl import load_workbook

import sys

import os
import numpy as np
import shutil
import re

from pybdv import transformations as tf
import mobie.metadata as mm

import glob
import json
import xml.etree.ElementTree as ET

import mrcfile as mrc

def vec2mat(a,b):
    
    # returns a rotation matrix that trnasforms vector a on top of vector b
    
    
    a = a/np.linalg.norm(a)
    b = b/np.linalg.norm(b)
    
    s = np.linalg.norm(np.cross(a, b))
    
    c = np.dot(a, b)
    
    if len(a)>2:
        G = np.matrix([[c,-s,0],[s,c,0],[0,0,1]])
        F = np.matrix([a,(b-c*a)/np.linalg.norm(b-c*a),np.cross(b,a)]).T
        rotmat = F @ G @ np.linalg.inv(F)
    else:
        rotmat = np.matrix([[c,s],[-s,c]])
    
    return rotmat




def cropvals(v1,v2,ax, minwindow=0.3,pxs=0.1):
    
    c = np.mean([v1,v2],axis=0)
    
    vec = np.diff([[v1],[v2]],axis=0).squeeze()

    
    M = vec2mat(vec,ax)
    
    c1 = M@c * pxs

    windowsize = np.max([minwindow,pxs*np.linalg.norm(vec)])

    cropmin = np.array(c1 - windowsize/2).squeeze()
    cropmax = np.array(c1 + windowsize/2).squeeze()

    M1 = np.vstack([M,[[0,0,0]]])
    M1 = np.hstack([M1,[[0],[0],[0],[1]]])

    trafo = tf.matrix_to_transformation(M1)

    return trafo,cropmin,cropmax




# ===================================================

# Script starts here




# load Excel file

strain = sys.argv[1] #'B-CLL'

indir = './analysis'
datadir = './data/tomo/'
imagedir = os.path.join(datadir,'images','ome-zarr')

overwr_source = False

a=glob.glob(os.path.join(indir,strain,'*xlsx'))

# if len(a) > 1:
#     raise IndexError('more than one file found')

if len(a) < 1:
    raise FileNotFoundError('no Excel file found')    

joinfiles=[]
badfiles = False


for xlfile in a:
    skip=False

# xlfile = a[0]
    wb = load_workbook(xlfile)
    pxs = 0.0015578000068664551

    if not 'data' in wb.sheetnames: continue
    
    sheet = wb['data']
    
    # parse file ID

    # parse positions
    
    pcells = np.array(sheet['I:L'])[:,1:]
    
    x = [c.value for c in pcells[0]]
    y = [c.value for c in pcells[1]]
    z = [c.value for c in pcells[2]]
    
    c_id = [cell.value for cell in sheet['F'][1:]]

    t_id = [cell.value for cell in sheet['A'][1:]]

    # extract unique targets
    
    targets=[]

    filesin = glob.glob(os.path.join(imagedir,strain + '*'))


    for infile in np.unique(filesin):

        print('Adding centriole crops for tomogram ' + infile + '.\n......')
        
        grid = infile.split('_')[2]

        tomoid = infile.split('_')[3].split('.')[0]
        
        # # with open(newxml, 'r') as f: xmltxt = f.readlines()
        #
        # if os.path.exists(newxml):
        #
        #
        #     tree = ET.parse(newxml)
        #
        #     root = tree.getroot()
        #
        #     for child in root:
        #         if child.tag == 'SequenceDescription':
        #             for sqch in child:
        #                 if sqch.tag == 'ViewSetups':
        #                     for vsch in sqch[0]:
        #
        #                         if vsch.tag=='size':
        #                             imsize = vsch.text
        #
        #
        #     im_sz = list(map(int,imsize.split(' ')))
        #
        # # assemble join file path
        # filepath= tomoid + '.join'
        #
        # sourcename = newid+'_'+infile
        #
        #
        # if overwr_source:
        #
        #
        #
        #
        #     mrc_in = os.path.join(joindir,strain, patient,grid,filepath)
        #
        #     try:
        #         mfile = mrc.mmap(mrc_in,permissive = 'True')
        #     except:
        #         print('Error in '+mrc_in)
        #         badfiles = True
        #         continue
        #
        #     pxs = mfile.voxel_size.x / 10000 # in um
        #

        with open(os.path.join(infile, 's0','.zarray')) as f:
            im_md = json.load(f)

        im_sz = im_md['shape']


        #
        #     n5link = '_'+patient+'_'+infile
        #
        #     n5file = os.path.join(imagedir,strain,n5link+'.n5')
        #
        #
        #     newlink = os.path.join(strain,sourcename)
        #     newn5 = os.path.join(imagedir,newlink+'.n5')
        #
        #     if not os.path.exists(newn5):
        #         if not os.path.exists(n5file):
        #             print('adding '+infile+ ' to list of files to convert.')
        #             joinfiles.append(patient+'/'+infile)
        #             skip = True
        #             continue
        #
        #         # move n5
        #         shutil.move(n5file,newn5)
        #
        #
        #
        #     # change link in XML
        #     xmlfile = os.path.join(imagedir,strain,'_'+patient+'_'+infile+'.xml')
        #
        #     if not os.path.exists(xmlfile):
        #         xmlfile = os.path.join(xmldir,'_'+patient+'_'+infile+'.xml')
        #
        #
        #     with open(xmlfile, 'r') as f: xmltxt = f.read()
        #
        #
        #     xmltxt = xmltxt.replace(n5link,newlink)
        #     xmltxt = xmltxt.replace(tomoid+'<',sourcename+'<')
        #     #remove original file reference
        #
        #     xmltxt = re.sub('\<OriginalFile\>.*\<\/OriginalFile\>\\n','',xmltxt)
        #
        #     with open(newxml, 'w') as f: f.write(xmltxt)
        #
        #     # os.remove(xmlfile)
        #
        #
        #
        #
        #     # add source to dataset
        #
        #     # generate tomo view
        #
        #
        #
        #     disp = mm.view_metadata.get_image_display(sourcename,[sourcename])
        #
        #
        #     view = mm.get_view(names = [sourcename],
        #                         source_types = ['image'],
        #                         sources = [[sourcename]],
        #                         display_settings = [disp],
        #                         is_exclusive = True,
        #                         menu_name = 'Tomograms'
        #                         )
        #
        #     mm.add_source_to_dataset(dataset_folder = datadir,
        #                            source_type = 'image',
        #                            source_name = sourcename,
        #                            image_metadata_path = newxml,
        #                            view = view,
        #                            overwrite = overwr_source
        #                            )
        #
        #     del(mfile)
        #
        #
        #
        #
        # find corresponding rows in data
        filerows = [i for i,item in enumerate(t_id) if item == tomoid]
    
        filelabels = [c_id[i] for i in filerows]
    
        lenrows = [item for item in filerows if 'length' in c_id[item]]
    
        numcent = np.unique([c_id[i] for i in lenrows])

        for cent in numcent:
            
            
            prefix=''
    
            if len(numcent)>1: cent_idx = '_' + cent.split('.')[0]
    
    
            t=dict()
            lenpts = [filerows[i] for i,label in enumerate(filelabels) if (cent.partition('.')[0] in label) & ('length' in label)]
    
            lenpos = np.stack([np.array((x[pt], im_sz[1] - y[pt],z[pt])) for pt in lenpts])
    
            if lenpos.shape[0]<2:
                raise ValueError
    
            if lenpos.shape[0]>2:
                veclength = np.sum(np.square(np.diff(lenpos,axis=0)),axis=1)
                lenpos = lenpos[np.argmax(veclength):(np.argmax(veclength)+2),:]
    
            # midpts = [filerows[i] for i,label in enumerate(filelabels) if (cent.partition('.')[0] in label) & ('diam.mid' in label)]
    
            # midpos = np.array((x[midpts[0]],y[midpts[0]],z[midpts[0]]))

            # t['file']=filepath
            # t['lpt']=lenpos
            # t['midpt']=midpos

            sourcename = os.path.basename(infile).rstrip('.ome.zarr')
    
            trafo,cropmin,cropmax = cropvals(lenpos[0,:],lenpos[1,:],[1,0,0],minwindow=1,pxs=pxs)
            
            # targets.append(t)
            
            centname = sourcename + '_' + cent.split('.')[0]
            
            
            # generate individual centriole view
            
            affine_trafo = mm.get_affine_source_transform(sources = [sourcename],
                                                          parameters = trafo,
                                                          source_names_after_transform = [centname + '_tfm']
                                                          )
            
            # XXXXXXX    
            # # crop not yet implemented...
            
            # crop = {"crop":{"max":list(cropmax),
            #     "min":list(cropmin),
            #     "shiftToOrigin": True,
            #     "sourceNamesAfterTransform": [
            #         centname + "_crop"
            #         ],
            #     "sources" : [centname + '_tfm']
            #         }
            #     }
            
            
            crop = mm.get_crop_source_transform(sources = [centname + '_tfm'],
                                                min = cropmin,
                                                max = cropmax,                                                                                               
                                                source_names_after_transform = [
                                                    centname + "_crop"
                                                    ]
                                                )

            disp1 = mm.view_metadata.get_image_display(centname,[centname + "_crop"])      
            
            
            cview = mm.get_view(names = [centname],
                    source_types = ['image'],
                    sources = [[centname + "_crop"]],
                    display_settings = [disp1],
                    is_exclusive = True,
                    menu_name = 'Centrioles',
                    source_transforms = [affine_trafo,                                         
                                         crop])
            
            # # XXXXXXXXXXXXXX
            # # manual for now...            
            # cview['sourceTransforms'].append(crop)
            
        
            mm.add_view_to_dataset(dataset_folder = datadir,                                       
                                   view_name = centname + "_crop",
                                   view = cview)

    
if not badfiles:
    
    with open(strain+'_joinfiles.txt','w') as f:
       f.write('\n'.join(joinfiles))

            # normvec = np.cross(np.diff(lenpos,axis=0),lenpos[0]-midpos)
            # normvec = normvec/np.linalg.norm(normvec)
            # v1 = np.cross(normvec,[0,0,1])
            # v1 = v1/np.linalg.norm(v1)
            # v2 = np.cross(normvec,v1)
            # v2 = v2/np.linalg.norm(v2)
